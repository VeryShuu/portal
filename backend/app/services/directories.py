"""Business logic for the object directories feature (docs/wip/directories.md).

HTTP-agnostic helpers: CRUD for directory *types*, *entries* and their
*contacts*, validation of ``attributes`` against the type ``field_schema`` and
of contact ``channel`` against the type ``channels``, name search, and export
builders (CSV / XLSX / HTML-for-PDF).
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import UTC, date, datetime
from html import escape
from typing import Any

from fastapi import HTTPException, status
from sqlalchemy import Select, case, func, or_, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.files import FileFolder
from app.models.object_directory import (
    ObjectDirectory,
    ObjectDirectoryEntry,
    ObjectEntryContact,
)
from app.schemas.object_directory import (
    ContactInput,
    CreateDirectoryRequest,
    CreateEntryRequest,
    DirectoryField,
    EntryReorderItem,
    UpdateDirectoryRequest,
    UpdateEntryRequest,
)


def _escape_like(q: str) -> str:
    """Escape LIKE/ILIKE wildcards so user input matches literally."""
    return q.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


# ── Validation ────────────────────────────────────────────────────────────────


def _parse_field_schema(raw: Any) -> list[DirectoryField]:
    """Coerce the stored JSONB ``field_schema`` into typed field definitions."""
    return [DirectoryField.model_validate(item) for item in (raw or [])]


def validate_attributes(field_schema: Any, attributes: dict[str, Any]) -> dict[str, str]:
    """Validate and normalize ``attributes`` against a directory ``field_schema``.

    - rejects keys not declared in the schema (422),
    - enforces ``required`` fields (422),
    - validates per-field ``type`` (number/email/url), and
    - normalizes every value to a trimmed string.
    """
    fields = _parse_field_schema(field_schema)
    allowed = {f.key: f for f in fields}

    unknown = set(attributes) - set(allowed)
    if unknown:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=f"Unknown attribute keys: {', '.join(sorted(unknown))}",
        )

    normalized: dict[str, str] = {}
    for key, field in allowed.items():
        raw = attributes.get(key)
        value = "" if raw is None else str(raw).strip()
        if not value:
            if field.required:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                    detail=f"Field '{key}' is required",
                )
            continue
        if field.type == "number":
            try:
                float(value.replace(",", "."))
            except ValueError as exc:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                    detail=f"Field '{key}' must be a number",
                ) from exc
        elif field.type == "email":
            if "@" not in value or value.startswith("@") or value.endswith("@"):
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                    detail=f"Field '{key}' must be a valid email",
                )
        elif field.type == "url" and not value.startswith(("http://", "https://")):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail=f"Field '{key}' must be an http/https URL",
            )
        normalized[key] = value
    return normalized


def validate_channels(channels: Any, contacts: list[ContactInput]) -> None:
    """Ensure every contact ``channel`` is declared in the directory ``channels``."""
    allowed = {item["key"] for item in (channels or []) if isinstance(item, dict) and "key" in item}
    for contact in contacts:
        if contact.channel not in allowed:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                detail=f"Unknown channel: {contact.channel}",
            )


# ── Directory (type) CRUD ──────────────────────────────────────────────────────


async def list_directories(
    db: AsyncSession, *, include_disabled: bool
) -> tuple[list[ObjectDirectory], int]:
    """Return active directory types ordered by ``sort_order`` then label."""
    conditions: list[Any] = [ObjectDirectory.deleted_at.is_(None)]
    if not include_disabled:
        conditions.append(ObjectDirectory.enabled.is_(True))
    stmt = (
        select(ObjectDirectory)
        .where(*conditions)
        .order_by(ObjectDirectory.sort_order, ObjectDirectory.label_ru)
    )
    items = list((await db.execute(stmt)).scalars().all())
    return items, len(items)


async def get_directory_or_404(db: AsyncSession, directory_id: uuid.UUID) -> ObjectDirectory:
    result = await db.execute(
        select(ObjectDirectory).where(
            ObjectDirectory.id == directory_id, ObjectDirectory.deleted_at.is_(None)
        )
    )
    directory = result.scalar_one_or_none()
    if not directory:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Directory not found")
    return directory


async def get_directory_by_slug_or_404(db: AsyncSession, slug: str) -> ObjectDirectory:
    result = await db.execute(
        select(ObjectDirectory).where(
            ObjectDirectory.slug == slug, ObjectDirectory.deleted_at.is_(None)
        )
    )
    directory = result.scalar_one_or_none()
    if not directory:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Directory not found")
    return directory


async def create_directory(
    db: AsyncSession, body: CreateDirectoryRequest
) -> ObjectDirectory:
    directory = ObjectDirectory(
        slug=body.slug,
        label_ru=body.label_ru,
        label_en=body.label_en,
        icon=body.icon,
        description=body.description,
        field_schema=[f.model_dump() for f in body.field_schema],
        channels=[c.model_dump() for c in body.channels],
        enabled=body.enabled,
        sort_order=body.sort_order,
    )
    db.add(directory)
    try:
        await db.commit()
    except IntegrityError as exc:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Directory slug already exists",
        ) from exc
    await db.refresh(directory)
    return directory


async def update_directory(
    db: AsyncSession, directory: ObjectDirectory, body: UpdateDirectoryRequest
) -> list[str]:
    changes = body.model_dump(exclude_unset=True)
    for field, value in changes.items():
        if field == "field_schema" and value is not None:
            directory.field_schema = [f.model_dump() for f in body.field_schema or []]
        elif field == "channels" and value is not None:
            directory.channels = [c.model_dump() for c in body.channels or []]
        else:
            setattr(directory, field, value)
    directory.updated_at = datetime.now(UTC)
    await db.commit()
    await db.refresh(directory)
    return sorted(changes.keys())


async def delete_directory(db: AsyncSession, directory: ObjectDirectory) -> None:
    """Soft-delete a directory type (its entries stay but the tab disappears)."""
    directory.deleted_at = datetime.now(UTC)
    await db.commit()


# ── Entry CRUD ────────────────────────────────────────────────────────────────

_ENTRY_LOAD = (
    selectinload(ObjectDirectoryEntry.contacts),
    selectinload(ObjectDirectoryEntry.folder),
)


async def _ensure_folder_exists(db: AsyncSession, folder_id: uuid.UUID | None) -> None:
    if folder_id is None:
        return
    exists = await db.scalar(
        select(FileFolder.id).where(
            FileFolder.id == folder_id,
            FileFolder.deleted_at.is_(None),
        )
    )
    if exists is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Folder not found",
        )


async def list_entries(
    db: AsyncSession,
    *,
    directory_id: uuid.UUID,
    q: str | None,
    limit: int,
    offset: int,
) -> tuple[list[ObjectDirectoryEntry], int]:
    conditions: list[Any] = [
        ObjectDirectoryEntry.directory_id == directory_id,
        ObjectDirectoryEntry.deleted_at.is_(None),
    ]
    if q:
        conditions.append(ObjectDirectoryEntry.name.ilike(f"%{_escape_like(q)}%", escape="\\"))

    count_stmt = select(func.count()).select_from(ObjectDirectoryEntry).where(*conditions)
    total: int = (await db.execute(count_stmt)).scalar_one()

    stmt = (
        select(ObjectDirectoryEntry)
        .where(*conditions)
        .options(*_ENTRY_LOAD)
        .order_by(ObjectDirectoryEntry.sort_order, ObjectDirectoryEntry.name)
        .offset(offset)
        .limit(limit)
    )
    items = list((await db.execute(stmt)).scalars().all())
    return items, total


async def get_entry_or_404(
    db: AsyncSession, *, directory_id: uuid.UUID, entry_id: uuid.UUID
) -> ObjectDirectoryEntry:
    result = await db.execute(
        select(ObjectDirectoryEntry)
        .where(
            ObjectDirectoryEntry.id == entry_id,
            ObjectDirectoryEntry.directory_id == directory_id,
            ObjectDirectoryEntry.deleted_at.is_(None),
        )
        .options(*_ENTRY_LOAD)
    )
    entry = result.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entry not found")
    return entry


def _build_contacts(contacts: list[ContactInput]) -> list[ObjectEntryContact]:
    return [
        ObjectEntryContact(
            role=c.role,
            channel=c.channel,
            label=c.label,
            value=c.value,
            sort_order=c.sort_order,
        )
        for c in contacts
    ]


async def create_entry(
    db: AsyncSession,
    *,
    directory: ObjectDirectory,
    body: CreateEntryRequest,
    created_by: uuid.UUID,
) -> ObjectDirectoryEntry:
    attributes = validate_attributes(directory.field_schema, body.attributes)
    validate_channels(directory.channels, body.contacts)
    await _ensure_folder_exists(db, body.folder_id)

    entry = ObjectDirectoryEntry(
        directory_id=directory.id,
        name=body.name,
        folder_id=body.folder_id,
        attributes=attributes,
        note=body.note,
        sort_order=body.sort_order,
        created_by=created_by,
        contacts=_build_contacts(body.contacts),
    )
    db.add(entry)
    await db.commit()
    return await get_entry_or_404(db, directory_id=directory.id, entry_id=entry.id)


async def update_entry(
    db: AsyncSession,
    *,
    directory: ObjectDirectory,
    entry: ObjectDirectoryEntry,
    body: UpdateEntryRequest,
) -> tuple[ObjectDirectoryEntry, list[str]]:
    changes = body.model_dump(exclude_unset=True)

    if "attributes" in changes and body.attributes is not None:
        entry.attributes = validate_attributes(directory.field_schema, body.attributes)
    if "name" in changes and body.name is not None:
        entry.name = body.name
    if "folder_id" in changes:
        await _ensure_folder_exists(db, body.folder_id)
        entry.folder_id = body.folder_id
    if "note" in changes:
        entry.note = body.note
    if "sort_order" in changes and body.sort_order is not None:
        entry.sort_order = body.sort_order
    if "contacts" in changes and body.contacts is not None:
        validate_channels(directory.channels, body.contacts)
        entry.contacts = _build_contacts(body.contacts)

    entry.updated_at = datetime.now(UTC)
    await db.commit()
    refreshed = await get_entry_or_404(db, directory_id=directory.id, entry_id=entry.id)
    return refreshed, sorted(changes.keys())


async def soft_delete_entry(db: AsyncSession, entry: ObjectDirectoryEntry) -> None:
    entry.deleted_at = datetime.now(UTC)
    await db.commit()


async def reorder_entries(
    db: AsyncSession,
    *,
    directory_id: uuid.UUID,
    items: list[EntryReorderItem],
) -> None:
    """Apply new ``sort_order`` to entries of one directory in a single update.

    Every id must reference an active (non-deleted) entry of ``directory_id``;
    otherwise a 404 is raised and nothing is changed.
    """
    if not items:
        return
    request_ids = {item.id for item in items}
    existing_result = await db.execute(
        select(ObjectDirectoryEntry.id).where(
            ObjectDirectoryEntry.directory_id == directory_id,
            ObjectDirectoryEntry.deleted_at.is_(None),
            ObjectDirectoryEntry.id.in_(list(request_ids)),
        )
    )
    existing_ids = {row[0] for row in existing_result.all()}
    if existing_ids != request_ids:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="One or more entries not found",
        )

    when_clauses = [(ObjectDirectoryEntry.id == item.id, item.sort_order) for item in items]
    sort_case = case(*when_clauses, else_=ObjectDirectoryEntry.sort_order)
    await db.execute(
        update(ObjectDirectoryEntry)
        .where(ObjectDirectoryEntry.id.in_(list(request_ids)))
        .values(sort_order=sort_case, updated_at=datetime.now(UTC))
    )
    await db.commit()


# ── Search (by name only) ──────────────────────────────────────────────────────


def directory_entry_search_stmt(q: str) -> Select[Any]:
    """Build a SELECT of (entry, directory_slug) for active entries matching ``q``.

    Only entries whose directory is enabled and not deleted are returned.
    """
    q_esc = _escape_like(q)
    return (
        select(ObjectDirectoryEntry, ObjectDirectory.slug, ObjectDirectory.label_ru)
        .join(ObjectDirectory, ObjectDirectoryEntry.directory_id == ObjectDirectory.id)
        .where(
            ObjectDirectoryEntry.deleted_at.is_(None),
            ObjectDirectory.deleted_at.is_(None),
            ObjectDirectory.enabled.is_(True),
            or_(ObjectDirectoryEntry.name.ilike(f"%{q_esc}%", escape="\\")),
        )
    )


# ── Export ────────────────────────────────────────────────────────────────────


def _format_contacts(entry: ObjectDirectoryEntry, channel_labels: dict[str, str]) -> str:
    parts: list[str] = []
    for c in sorted(entry.contacts, key=lambda x: x.sort_order):
        chan = channel_labels.get(c.channel, c.channel)
        prefix = f"{c.role} · " if c.role else ""
        parts.append(f"{prefix}{chan}: {c.value}")
    return "; ".join(parts)


def build_export_table(
    directory: ObjectDirectory, entries: list[ObjectDirectoryEntry]
) -> tuple[list[str], list[list[str]]]:
    """Return ``(headers, rows)`` for the given directory and its entries."""
    fields = sorted(_parse_field_schema(directory.field_schema), key=lambda f: f.sort_order)
    channel_labels: dict[str, str] = {
        str(item["key"]): str(item.get("label_ru", item["key"]))
        for item in (directory.channels or [])
        if isinstance(item, dict) and "key" in item
    }

    headers = ["Название", *[f.label_ru for f in fields], "Контакты", "Заметка"]
    rows: list[list[str]] = []
    for entry in entries:
        row = [entry.name]
        row.extend(str(entry.attributes.get(f.key, "")) for f in fields)
        row.append(_format_contacts(entry, channel_labels))
        row.append(entry.note or "")
        rows.append(row)
    return headers, rows


def export_filename(directory: ObjectDirectory, ext: str) -> str:
    return f"{directory.slug}-{date.today().isoformat()}.{ext}"


def build_csv(directory: ObjectDirectory, entries: list[ObjectDirectoryEntry]) -> bytes:
    headers, rows = build_export_table(directory, entries)
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(headers)
    writer.writerows(rows)
    # BOM so Excel opens UTF-8 CSV with Cyrillic correctly.
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def build_xlsx(directory: ObjectDirectory, entries: list[ObjectDirectoryEntry]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers, rows = build_export_table(directory, entries)

    wb = Workbook()
    ws = wb.active
    ws.title = directory.label_ru[:31] or "Справочник"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = 28 if idx == len(headers) - 1 else 22

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_export_html(directory: ObjectDirectory, entries: list[ObjectDirectoryEntry]) -> str:
    headers, rows = build_export_table(directory, entries)
    head_html = "".join(f"<th>{escape(h)}</th>" for h in headers)
    body_html = "".join(
        "<tr>" + "".join(f"<td>{escape(cell)}</td>" for cell in row) + "</tr>" for row in rows
    )
    title = escape(directory.label_ru)
    generated = date.today().strftime("%d.%m.%Y")
    return (
        "<!DOCTYPE html><html lang='ru'><head><meta charset='utf-8'>"
        "<style>"
        "body{font-family:Arial,Helvetica,sans-serif;font-size:11px;color:#1a1a1a;}"
        "h1{font-size:16px;margin:0 0 2px;}"
        ".sub{color:#666;font-size:10px;margin-bottom:10px;}"
        "table{width:100%;border-collapse:collapse;}"
        "th,td{border:1px solid #bfbfbf;padding:4px 6px;text-align:left;"
        "vertical-align:top;word-break:break-word;}"
        "th{background:#305496;color:#fff;}"
        "tr:nth-child(even) td{background:#f2f5fb;}"
        "</style></head><body>"
        f"<h1>{title}</h1><div class='sub'>Сформирован: {generated}</div>"
        f"<table><thead><tr>{head_html}</tr></thead><tbody>{body_html}</tbody></table>"
        "</body></html>"
    )

"""Общие helpers для XLSX-экспортов: стилизация шапки таблицы.

Вынесено из ``services/directories.build_xlsx`` и ``api/analytics.py``
(см. audit [L9]): при ребрендинге (акцентный цвет, шрифт) правка идёт в одном
месте, а не в нескольких модулях одновременно.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Корпоративная палитра экспорта. TODO(owner: branding) — вынести в SystemSettings,
# если брендинг станет per-deployment. Сейчас — константы проекта.
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(
    ws: Worksheet,
    *,
    headers: list[str],
    column_widths: int | list[int] = 22,
) -> None:
    """Применить корпоративный стиль к шапке XLSX-таблицы.

    * Bold + белый шрифт на синем фоне (``HEADER_FONT``/``HEADER_FILL``);
    * Центрирование + wrap_text;
    * column_widths — либо одно число для всех колонок, либо список длин
      (для особых случаев вроде directories, где последняя колонка шире).
    """
    widths = [column_widths] * len(headers) if isinstance(column_widths, int) else column_widths

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1]


def fill_body_cells(ws: Worksheet, *, ncols: int, start_row: int = 2) -> None:
    """Применить ``BODY_ALIGNMENT`` ко всем ячейкам тела таблицы.

    Используется в ``directories.build_xlsx`` для единообразного wrap_text.
    """
    for row in ws.iter_rows(min_row=start_row, max_col=ncols):
        for cell in row:
            cell.alignment = BODY_ALIGNMENT


def freeze_header(ws: Worksheet) -> None:
    """Закрепить шапку (строка 1) при скролле — idempotent."""
    ws.freeze_panes = "A2"

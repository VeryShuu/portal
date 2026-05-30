"""XLSX-генератор печатной версии справочника сотрудников."""

from __future__ import annotations

import io
from datetime import date

from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.utils.phone import apply_phone_regex

from . import users_repo

_HEADERS = ["№", "Ф.И.О.", "Должность", "Внутр.", "Мобильный", "E-mail", "Город"]
_COL_WIDTHS = [5, 36, 44, 10, 18, 30, 16]
_TITLE = 'Справочник сотрудников АО "МАГЭ"'

_THIN_SIDE = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_TITLE_FONT = Font(name="Calibri", size=14, bold=True)
_SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_DEPT_FONT = Font(name="Calibri", size=11, bold=True)
_DEPT_FILL = PatternFill("solid", fgColor="D9E1F2")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

_HEADER_ROW = 4


async def export_users_xlsx(
    db: AsyncSession,
    *,
    q: str | None,
    department: str | None,
    office: str | None,
    sort: str,
    phone_regex: str,
) -> Response:
    """Собрать XLSX-справочник и вернуть FastAPI Response (attachment)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Справочник"

    n_cols = len(_HEADERS)
    last_col = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = _TITLE
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Сформирован: {date.today().strftime('%d.%m.%Y')}"
    ws["A2"].font = _SUBTITLE_FONT
    ws["A2"].alignment = _CENTER

    for idx, h in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[_HEADER_ROW].height = 22

    for idx, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = f"A{_HEADER_ROW + 1}"

    row = _HEADER_ROW + 1
    current_dept: str | None = None
    seq_in_dept = 0

    async for user in users_repo.stream_users(
        db,
        q=q,
        department=department,
        office=office,
        sort=sort,
        include_hidden=(sort != "staff_custom"),
    ):
        dept = (user.department or "").strip() or "—"
        if dept != current_dept:
            current_dept = dept
            seq_in_dept = 0
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
            cell = ws.cell(row=row, column=1, value=dept)
            cell.font = _DEPT_FONT
            cell.fill = _DEPT_FILL
            cell.alignment = _LEFT_WRAP
            for c in range(1, n_cols + 1):
                ws.cell(row=row, column=c).border = _BORDER
            row += 1

        seq_in_dept += 1
        attrs = user.attributes or {}
        values = [
            seq_in_dept,
            user.full_name or "",
            user.position or "",
            apply_phone_regex(user.phone or "", phone_regex),
            attrs.get("mobile", "") or "",
            user.email or "",
            attrs.get("city", "") or "",
        ]
        for idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=idx, value=value)
            cell.border = _BORDER
            cell.alignment = _CENTER if idx in (1, 4) else _LEFT_WRAP
        row += 1

    ws.print_title_rows = f"{_HEADER_ROW}:{_HEADER_ROW}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"staff-{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store, max-age=0",
        },
    )
